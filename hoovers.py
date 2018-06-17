file_name = 'Revenue Lookup by DUNS.xlsx'
import traceback, time
from itertools import count

from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook

DEBUG = False
if DEBUG:
	import requests_cache
	requests_cache.install_cache(allowable_methods=('GET', 'POST'))

class Crawler:
	def __init__(self):
		self.session = requests.Session()
		self.session.headers.update((
			('User-Agent', 'Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.94 Safari/537.36'),
			('Accept', '*/*'),
			('Accept-Language', 'en-US,en;q=0.8'),
		))

	def get(self, url, data=None):
		for i_try in count(1):
			try:
				if data:
					resp = self.session.post(url, data=data, verify=False, timeout=30)
				else:
					resp = self.session.get(url, verify=False, timeout=30)
				if resp.status_code == 500:
					print(url)
					print('server error')
					return None
				resp.raise_for_status()
				return BeautifulSoup(resp.text, 'lxml')
			except requests.RequestException as e:
				print(url)
				print(e)
			time.sleep(min(2**i_try, 3600))

	def main(self):
		wb = load_workbook(file_name)
		ws = wb.active
		try:
			for row in range(2, ws.max_row + 1):
				try:
					url = ws.cell(row=row, column=3).value
					if not url:
						continue
					soup = self.get(url)
					if soup is not None:
						trs = soup.select('div.data-table > table > tbody > tr')
						if trs:
							value = None
							cnt_us = 0
							for tr in trs:
								tds = tr('td', recursive=False, limit=3)
								v = tds[2].get_text(strip=True)
								if ', United States' in tds[1].string:
									if cnt_us == 0 or not value:
										value = v
									cnt_us += 1
								elif not cnt_us and not value:
									value = v
								ws.cell(row=row, column=4, value=value)
							assert len(trs) == 1 or cnt_us == 1, url
				except Exception:
					print('row', row)
					traceback.print_exc()
				if row % 100 == 0:
					print('row', row)
					wb.save(file_name)
		finally:
			wb.save(file_name)

if __name__ == '__main__':
	requests.packages.urllib3.disable_warnings()
	Crawler().main()
