file_name = 'NASA Scrape.xlsx'
import traceback, time
from itertools import count

from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook

DEBUG = False
if DEBUG:
	import requests_cache
	requests_cache.install_cache()

class Crawler:
	def __init__(self):
		self.session = requests.Session()
		self.session.headers.update((
			('User-Agent', 'Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.94 Safari/537.36'),
			('Accept', '*/*'),
			('Accept-Language', 'en-US,en;q=0.8'),
		))
		self.url = None

	def get(self, url, data=None):
		self.url = url
		for i_try in count(1):
			try:
				if data:
					resp = self.session.post(url, data=data, verify=False, timeout=30)
				else:
					resp = self.session.get(url, verify=False, timeout=30)
				resp.raise_for_status()
				return BeautifulSoup(resp.text, 'lxml')
			except requests.RequestException as e:
				print(url)
				print(e)
			time.sleep(min(2**i_try, 3600))

	def main(self):
		wb = load_workbook(file_name)
		ws = wb.active
		for row in range(2, ws.max_row + 1):
			url = ws.cell(row=row, column=1).value
			if not url:
				continue
			soup = self.get(url)
			p = soup.find('span', string='Lead Center:').parent.previous_sibling
			t = []
			for el in p.next_siblings:
				process(el, t)
				t.append('\n\n')
			t = ''.join(t).strip()
			with open(url[-5:] + '.txt', 'w', encoding='utf_8') as f: f.write(t)
			ws.cell(row=row, column=2, value=t)
		wb.save(file_name)

def process(el, t):
	for el in el.children:
		if isinstance(el, str):
			if not el.isspace():
				t.append(el.strip('\n').replace('\xa0', ' '))
		elif el.name in ('span', 'strong') and not (el.string or '').isspace() or (el.name == 'em' and el.string == 'References:'):
			if not t or t[-1].endswith('\n'):
				t.append('#### ')
			process(el, t)
		elif el.name in ('p', 'div'):
			process(el, t)
			t.append('\n\n')
		elif el.name == 'a' and el.string in ('Read more>>', 'Read less>>'):
			pass
		elif el.name == 'li':
			level = len(el.find_parents(['ul', 'blockquote'])) - 2 # have extra ul on top
			t.append('    '*level + '* ')
			process(el, t)
			t.append('\n')
		elif el.name == 'br':
			if el.next_sibling:
				t.append('\n')
		elif el.name == 'ul':
			process(el, t)
			t.append('\n')
		else:
			if el.name not in ('blockquote', 'sub', 'sup', 'a', 'em', 'span', 'strong'):
				raise RuntimeError(el.name)
			process(el, t)

if __name__ == '__main__':
	requests.packages.urllib3.disable_warnings()
	try:
		c = Crawler()
		c.main()
	except Exception:
		print()
		print(c.url)
		raise
